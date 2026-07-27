"""
export_excel.py - Excel Workbook Generator (Multi-Domain)

Produces the final Demographics_Data.xlsx workbook with:
- Domain separator sheets (Demographics & Census, Economy & Employment)
- One worksheet per dataset with metadata header rows
- Dataset_Catalog, Data_Dictionary, Quality_Report worksheets
- Professional formatting (frozen headers, auto-filter, auto-sized columns)
"""

import logging
from datetime import datetime
from pathlib import Path

import pandas as pd
import numpy as np

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

logger = logging.getLogger(__name__)

PROCESSED_DIR = Path(__file__).parent.parent / "processed"
PROCESSED_DIR.mkdir(parents=True, exist_ok=True)
EXCEL_OUTPUT = PROCESSED_DIR / "Demographics_Data.xlsx"

# Styling
HEADER_FONT = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
THIN_BORDER = Border(
    left=Side(style="thin", color="B4C6E7"),
    right=Side(style="thin", color="B4C6E7"),
    top=Side(style="thin", color="B4C6E7"),
    bottom=Side(style="thin", color="B4C6E7"),
)
METADATA_FONT = Font(name="Calibri", size=10, italic=True)
METADATA_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
METADATA_LABEL_FONT = Font(name="Calibri", size=10, bold=True, color="1F3864")
ALT_FILL = PatternFill(start_color="F2F7FC", end_color="F2F7FC", fill_type="solid")
DATA_FONT = Font(name="Calibri", size=10)

# Domain separator styling
DOMAIN_FONT = Font(name="Calibri", bold=True, size=14, color="FFFFFF")
DOMAIN_FILL = PatternFill(start_color="1F3864", end_color="1F3864", fill_type="solid")

DATE_STR = datetime.now().strftime("%Y-%m-%d")


# ── Domain Metadata ───────────────────────────────────────────────────────────

DOMAINS = {
    "Demographics & Census": {
        "description": "Population, age, race, migration, households, counties, cities",
        "agencies": "U.S. Census Bureau",
        "color": "2F5496",
    },
    "Economy & Employment": {
        "description": "GDP, income, inflation, unemployment, wages, business statistics",
        "agencies": "Bureau of Economic Analysis (BEA), Bureau of Labor Statistics (BLS)",
        "color": "548235",
    },
}

# Dataset metadata for both domains
ALL_METADATA = {
    # Demographics
    "PEP_County": {"domain": "Demographics & Census", "Dataset Name": "PEP County Population Estimates", "Agency": "U.S. Census Bureau",
                   "Source URL": "https://www2.census.gov/programs-surveys/popest/datasets/",
                   "Download Date": DATE_STR, "Years Covered": "2020-2023",
                   "Update Frequency": "Annual", "License": "Public Domain",
                   "Geographic Level": "County", "Last Updated": DATE_STR,
                   "Notes": "Annual population estimates by county from PEP"},
    "Population_Estimates": {"domain": "Demographics & Census", "Dataset Name": "PEP State Population Estimates", "Agency": "U.S. Census Bureau",
                   "Source URL": "https://www2.census.gov/programs-surveys/popest/datasets/",
                   "Download Date": DATE_STR, "Years Covered": "2020-2023",
                   "Update Frequency": "Annual", "License": "Public Domain",
                   "Geographic Level": "State", "Last Updated": DATE_STR,
                   "Notes": "Annual population estimates by state from PEP"},
    "Population": {"domain": "Demographics & Census", "Dataset Name": "Population - ACS", "Agency": "U.S. Census Bureau",
                   "Source URL": "https://api.census.gov", "Download Date": DATE_STR,
                   "Years Covered": "2023", "Update Frequency": "Annual",
                   "License": "Public Domain", "Geographic Level": "State, County",
                   "Last Updated": DATE_STR, "Notes": "Total population from ACS"},
    "State_Lookup": {"domain": "Demographics & Census", "Dataset Name": "State FIPS Lookup", "Agency": "U.S. Census Bureau",
                   "Source URL": "N/A", "Download Date": DATE_STR, "Years Covered": "Current",
                   "Update Frequency": "As needed", "License": "Public Domain",
                   "Geographic Level": "State", "Last Updated": DATE_STR,
                   "Notes": "FIPS codes, state names, abbreviations"},
    "County_Lookup": {"domain": "Demographics & Census", "Dataset Name": "County FIPS Lookup", "Agency": "U.S. Census Bureau",
                   "Source URL": "N/A", "Download Date": DATE_STR, "Years Covered": "2023",
                   "Update Frequency": "Annual", "License": "Public Domain",
                   "Geographic Level": "County", "Last Updated": DATE_STR,
                   "Notes": "State FIPS, County FIPS, county names"},
    "ACS_Geography": {"domain": "Demographics & Census", "Dataset Name": "ACS Geography", "Agency": "U.S. Census Bureau",
                   "Source URL": "https://www2.census.gov/programs-surveys/acs/summary_file/2023/data/",
                   "Download Date": DATE_STR, "Years Covered": "2023",
                   "Update Frequency": "Annual", "License": "Public Domain",
                   "Geographic Level": "All", "Last Updated": DATE_STR,
                   "Notes": "Geographic header file for ACS 1-Year estimates"},
    # Economy
    "BEA_GDP": {"domain": "Economy & Employment", "Dataset Name": "Gross Domestic Product (NIPA)", "Agency": "Bureau of Economic Analysis",
                "Source URL": "https://apps.bea.gov/national/Release/TXT/xls/", "Download Date": DATE_STR,
                "Years Covered": "Historical", "Update Frequency": "Quarterly/Annual",
                "License": "Public Domain", "Geographic Level": "National",
                "Last Updated": DATE_STR, "Notes": "NIPA Table 1.1.5 - GDP by component"},
    "BEA_Personal_Income": {"domain": "Economy & Employment", "Dataset Name": "Personal Income (NIPA)", "Agency": "Bureau of Economic Analysis",
                "Source URL": "https://apps.bea.gov/national/Release/TXT/xls/", "Download Date": DATE_STR,
                "Years Covered": "Historical", "Update Frequency": "Quarterly/Annual",
                "License": "Public Domain", "Geographic Level": "National",
                "Last Updated": DATE_STR, "Notes": "NIPA Table 2.1 - Personal Income and Disposition"},
    "BEA_Real_GDP": {"domain": "Economy & Employment", "Dataset Name": "Real GDP (Chained Dollars)", "Agency": "Bureau of Economic Analysis",
                "Source URL": "https://apps.bea.gov/national/Release/TXT/xls/", "Download Date": DATE_STR,
                "Years Covered": "Historical", "Update Frequency": "Quarterly/Annual",
                "License": "Public Domain", "Geographic Level": "National",
                "Last Updated": DATE_STR, "Notes": "NIPA Table 1.1.6 - Real GDP, chained dollars"},
    "BEA_State_GDP": {"domain": "Economy & Employment", "Dataset Name": "State GDP by Area", "Agency": "Bureau of Economic Analysis",
                "Source URL": "https://apps.bea.gov/regional/xls/", "Download Date": DATE_STR,
                "Years Covered": "Historical", "Update Frequency": "Annual",
                "License": "Public Domain", "Geographic Level": "State, MSA, County",
                "Last Updated": DATE_STR, "Notes": "GDP by state and metropolitan area"},
    "BEA_State_Income": {"domain": "Economy & Employment", "Dataset Name": "State Personal Income", "Agency": "Bureau of Economic Analysis",
                "Source URL": "https://apps.bea.gov/regional/xls/", "Download Date": DATE_STR,
                "Years Covered": "Historical", "Update Frequency": "Annual",
                "License": "Public Domain", "Geographic Level": "State, County",
                "Last Updated": DATE_STR, "Notes": "Personal income by state and county"},
    "BEA_GDP_Industry": {"domain": "Economy & Employment", "Dataset Name": "GDP by Industry", "Agency": "Bureau of Economic Analysis",
                "Source URL": "https://apps.bea.gov/national/Release/TXT/xls/", "Download Date": DATE_STR,
                "Years Covered": "Historical", "Update Frequency": "Annual",
                "License": "Public Domain", "Geographic Level": "National",
                "Last Updated": DATE_STR, "Notes": "Value added by industry sector"},
    "BLS_LAUS_States": {"domain": "Economy & Employment", "Dataset Name": "State Unemployment (LAUS)", "Agency": "Bureau of Labor Statistics",
                "Source URL": "https://download.bls.gov/pub/time.series/la/", "Download Date": DATE_STR,
                "Years Covered": "1976-Present", "Update Frequency": "Monthly",
                "License": "Public Domain", "Geographic Level": "State",
                "Last Updated": DATE_STR, "Notes": "Local Area Unemployment Statistics, not seasonally adjusted"},
    "BLS_LAUS_States_SA": {"domain": "Economy & Employment", "Dataset Name": "State Unemployment (Seasonally Adjusted)", "Agency": "Bureau of Labor Statistics",
                "Source URL": "https://download.bls.gov/pub/time.series/la/", "Download Date": DATE_STR,
                "Years Covered": "1976-Present", "Update Frequency": "Monthly",
                "License": "Public Domain", "Geographic Level": "State",
                "Last Updated": DATE_STR, "Notes": "LAUS seasonally adjusted state data"},
    "BLS_CPI": {"domain": "Economy & Employment", "Dataset Name": "Consumer Price Index (CPI)", "Agency": "Bureau of Labor Statistics",
                "Source URL": "https://download.bls.gov/pub/time.series/cu/", "Download Date": DATE_STR,
                "Years Covered": "1947-Present", "Update Frequency": "Monthly",
                "License": "Public Domain", "Geographic Level": "National",
                "Last Updated": DATE_STR, "Notes": "CPI for all urban consumers, all items"},
    "BLS_CES": {"domain": "Economy & Employment", "Dataset Name": "Current Employment Statistics", "Agency": "Bureau of Labor Statistics",
                "Source URL": "https://download.bls.gov/pub/time.series/ce/", "Download Date": DATE_STR,
                "Years Covered": "1939-Present", "Update Frequency": "Monthly",
                "License": "Public Domain", "Geographic Level": "National",
                "Last Updated": DATE_STR, "Notes": "Nonfarm payroll employment, hours, earnings"},
}

# ACS table metadata (auto-generated)
for tbl in ["b01001","b01002","b02001","b03001","b03002","b07003","b11001","b11003",
            "b15003","b16001","b18101","b19001","b19013","b21001",
            "b25001","b25002","b25003","b25004","b25064","b25077"]:
    ALL_METADATA[f"ACS_{tbl}"] = {
        "domain": "Demographics & Census",
        "Dataset Name": f"ACS 1-Year {tbl.upper()} Estimates",
        "Agency": "U.S. Census Bureau",
        "Source URL": "https://www2.census.gov/programs-surveys/acs/summary_table/2023/data/",
        "Download Date": DATE_STR, "Years Covered": "2023",
        "Update Frequency": "Annual (1-Year)", "License": "Public Domain",
        "Geographic Level": "State, County, Place, Tract, CBSA, ZCTA",
        "Last Updated": DATE_STR, "Notes": f"ACS Table {tbl.upper()} from summary file",
    }

# Gazetteer metadata (auto-generated)
for gaz in ["Aiannh","Cbsa","County","Cousub","Elsd","Place","Scsd","Sldl","Sldu",
            "State","Tracts","Ua","Unsd","Zcta"]:
    ALL_METADATA[f"Gazetteer_{gaz}"] = {
        "domain": "Demographics & Census",
        "Dataset Name": f"Census Gazetteer - {gaz} File",
        "Agency": "U.S. Census Bureau",
        "Source URL": "https://www2.census.gov/geo/docs/maps-data/data/gazetteer/",
        "Download Date": DATE_STR, "Years Covered": "2024",
        "Update Frequency": "Annual", "License": "Public Domain",
        "Geographic Level": gaz, "Last Updated": DATE_STR,
        "Notes": f"Geographic reference data with land/water area for {gaz}",
    }


# ── Formatting Helpers ────────────────────────────────────────────────────────

def _auto_size_columns(ws, max_width=35, sample_rows=30):
    max_scan = min(ws.max_row + 1, sample_rows + 12)
    for col in range(1, min(ws.max_column + 1, 200)):
        max_len = 10
        col_letter = get_column_letter(col)
        for row in range(1, max_scan):
            val = ws.cell(row=row, column=col).value
            if val is not None:
                cl = len(str(val))
                if cl > max_len:
                    max_len = min(cl + 2, max_width)
        ws.column_dimensions[col_letter].width = max_len


def _write_domain_separator(wb, domain_name, domain_info):
    """Write a domain separator sheet."""
    ws = wb.create_sheet(title=f"-- {domain_name[:28]} --")
    ws.sheet_properties.tabColor = domain_info["color"]
    ws.cell(row=1, column=1, value=domain_name)
    ws.cell(row=1, column=1).font = DOMAIN_FONT
    ws.cell(row=1, column=1).fill = PatternFill(start_color=domain_info["color"], end_color=domain_info["color"], fill_type="solid")
    ws.merge_cells("A1:F1")
    ws.cell(row=2, column=1, value=domain_info["description"])
    ws.cell(row=2, column=1).font = Font(name="Calibri", size=11, italic=True, color="333333")
    ws.cell(row=3, column=1, value=f"Agencies: {domain_info['agencies']}")
    ws.cell(row=3, column=1).font = Font(name="Calibri", size=10, color="666666")
    return ws


def _write_dataset_sheet(wb, sheet_name, df, metadata):
    ws = wb.create_sheet(title=sheet_name[:31])

    if df.empty:
        ws.cell(row=1, column=1, value="No data available")
        return

    meta_keys = [
        ("Dataset Name", metadata.get("Dataset Name", "")),
        ("Agency", metadata.get("Agency", "")),
        ("Source URL", metadata.get("Source URL", "")),
        ("Download Date", metadata.get("Download Date", "")),
        ("Years Covered", metadata.get("Years Covered", "")),
        ("Update Frequency", metadata.get("Update Frequency", "")),
        ("License", metadata.get("License", "")),
        ("Geographic Level", metadata.get("Geographic Level", "")),
        ("Last Updated", metadata.get("Last Updated", "")),
        ("Notes", metadata.get("Notes", "")),
    ]

    for i, (label, value) in enumerate(meta_keys, start=1):
        c1 = ws.cell(row=i, column=1, value=label)
        c1.font = METADATA_LABEL_FONT
        c1.fill = METADATA_FILL
        c2 = ws.cell(row=i, column=2, value=str(value))
        c2.font = METADATA_FONT
        c2.fill = METADATA_FILL
        ws.merge_cells(start_row=i, start_column=2, end_row=i, end_column=10)

    header_row = 12
    cols = list(df.columns)

    for ci, cname in enumerate(cols, 1):
        ws.cell(row=header_row, column=ci, value=cname)
    for ci in range(1, len(cols) + 1):
        cell = ws.cell(row=header_row, column=ci)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN
        cell.border = THIN_BORDER

    data_start = header_row + 1
    use_bulk = len(df) > 2000

    for ri, (_, row) in enumerate(df.iterrows()):
        excel_row = data_start + ri
        for ci, val in enumerate(row):
            cell = ws.cell(row=excel_row, column=ci + 1)
            if pd.isna(val):
                cell.value = None
            elif isinstance(val, (np.integer,)):
                cell.value = int(val)
            elif isinstance(val, (np.floating,)):
                cell.value = float(val)
            else:
                cell.value = val
            if not use_bulk:
                cell.font = DATA_FONT
                cell.border = THIN_BORDER

    if not use_bulk:
        for ri in range(data_start, data_start + len(df)):
            if (ri - data_start) % 2 == 1:
                for ci in range(1, len(cols) + 1):
                    c = ws.cell(row=ri, column=ci)
                    if c.fill == PatternFill():
                        c.fill = ALT_FILL

    _auto_size_columns(ws)
    ws.freeze_panes = f"A{header_row + 1}"
    if len(df) > 0:
        last_col = get_column_letter(min(len(cols), 100))
        end_row = data_start + min(len(df), 5000) - 1
        ws.auto_filter.ref = f"A{header_row}:{last_col}{end_row}"


# ── Catalog & Dictionary ──────────────────────────────────────────────────────

def _write_catalog(wb, datasets):
    ws = wb.create_sheet(title="Dataset_Catalog")
    headers = ["Domain", "Dataset", "Description", "Agency", "Years", "Rows", "Columns",
               "Source URL", "Download Date", "Refresh Frequency"]

    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN
        cell.border = THIN_BORDER

    for ri, (name, df) in enumerate(datasets.items(), 2):
        m = ALL_METADATA.get(name, {})
        vals = [m.get("Domain", "Other"), name, m.get("Dataset Name", name),
                m.get("Agency", "U.S. Government"), m.get("Years Covered", "N/A"),
                len(df), len(df.columns), m.get("Source URL", ""),
                DATE_STR, m.get("Update Frequency", "Annual")]
        for ci, v in enumerate(vals, 1):
            ws.cell(row=ri, column=ci, value=v)
            ws.cell(row=ri, column=ci).font = DATA_FONT

    _auto_size_columns(ws)
    ws.freeze_panes = "A2"


def _write_dictionary(wb, datasets):
    ws = wb.create_sheet(title="Data_Dictionary")
    headers = ["Domain", "Worksheet", "Column Name", "Data Type", "Primary Key", "Nullable"]

    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN
        cell.border = THIN_BORDER

    key_cols = {"state", "county", "geoid", "geo_id", "fips", "tract", "place",
                "cbsa", "zcta", "us", "name", "fips_code", "state_fips", "county_fips",
                "series_id", "year", "period", "area_code", "measure_code"}
    type_map = {"int64": "Integer", "float64": "Float", "object": "String", "bool": "Boolean"}

    row_idx = 2
    for ws_name, df in datasets.items():
        domain = ALL_METADATA.get(ws_name, {}).get("Domain", "Other")
        nulls = df.isnull().any()
        dtypes = df.dtypes
        for col_name in df.columns:
            ws.cell(row=row_idx, column=1, value=domain)
            ws.cell(row=row_idx, column=2, value=ws_name)
            ws.cell(row=row_idx, column=3, value=col_name)
            dtype_str = str(dtypes[col_name])
            ws.cell(row=row_idx, column=4, value=type_map.get(dtype_str, dtype_str))
            ws.cell(row=row_idx, column=5, value="Yes" if col_name.lower() in key_cols else "")
            ws.cell(row=row_idx, column=6, value="Yes" if nulls[col_name] else "No")
            row_idx += 1

    ws.freeze_panes = "A2"
    if row_idx > 2:
        last_col = get_column_letter(len(headers))
        ws.auto_filter.ref = f"A1:{last_col}{row_idx - 1}"


def _write_quality(wb, datasets):
    ws = wb.create_sheet(title="Quality_Report")
    headers = ["Domain", "Dataset", "Rows", "Columns", "Duplicates", "Missing Cols",
               "Total Missing", "Date"]

    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN
        cell.border = THIN_BORDER

    for ri, (name, df) in enumerate(datasets.items(), 2):
        domain = ALL_METADATA.get(name, {}).get("Domain", "Other")
        dups = int(df.duplicated().sum())
        miss_cols = int(df.isnull().any().sum())
        miss_total = int(df.isnull().sum().sum())
        vals = [domain, name, len(df), len(df.columns), dups, miss_cols, miss_total, DATE_STR]
        for ci, v in enumerate(vals, 1):
            ws.cell(row=ri, column=ci, value=v)
            ws.cell(row=ri, column=ci).font = DATA_FONT

    _auto_size_columns(ws)
    ws.freeze_panes = "A2"


# ── Main Export ───────────────────────────────────────────────────────────────

def export_to_excel(datasets, validation_reports=None, output_path=None):
    if not HAS_OPENPYXL:
        raise ImportError("openpyxl required: pip install openpyxl")

    if output_path is None:
        output_path = EXCEL_OUTPUT

    logger.info("=" * 60)
    logger.info("EXCEL WORKBOOK GENERATION (MULTI-DOMAIN)")
    logger.info(f"Output: {output_path}")
    logger.info(f"Datasets: {len(datasets)}")
    logger.info("=" * 60)

    wb = Workbook()
    wb.remove(wb.active)

    # Group datasets by domain
    domain_datasets = {}
    for name, df in datasets.items():
        domain = ALL_METADATA.get(name, {}).get("Domain", "Other")
        if domain not in domain_datasets:
            domain_datasets[domain] = {}
        domain_datasets[domain][name] = df

    # Write each domain section
    for domain_name, domain_info in DOMAINS.items():
        if domain_name in domain_datasets:
            _write_domain_separator(wb, domain_name, domain_info)
            for name, df in domain_datasets[domain_name].items():
                meta = ALL_METADATA.get(name, {
                    "Dataset Name": name, "Agency": "U.S. Government",
                    "Source URL": "", "Download Date": DATE_STR, "Years Covered": "Various",
                    "Update Frequency": "Annual", "License": "Public Domain",
                    "Geographic Level": "Various", "Last Updated": DATE_STR, "Notes": "",
                })
                logger.info(f"Writing worksheet: {name} ({len(df)} rows)")
                _write_dataset_sheet(wb, name, df, meta)

    # Write remaining datasets (not in defined domains)
    for name, df in datasets.items():
        domain = ALL_METADATA.get(name, {}).get("Domain", "Other")
        if domain not in DOMAINS:
            meta = ALL_METADATA.get(name, {
                "Dataset Name": name, "Agency": "U.S. Government",
                "Source URL": "", "Download Date": DATE_STR, "Years Covered": "Various",
                "Update Frequency": "Annual", "License": "Public Domain",
                "Geographic Level": "Various", "Last Updated": DATE_STR, "Notes": "",
            })
            logger.info(f"Writing worksheet: {name} ({len(df)} rows)")
            _write_dataset_sheet(wb, name, df, meta)

    # Write summary sheets
    logger.info("Writing Dataset_Catalog")
    _write_catalog(wb, datasets)

    logger.info("Writing Data_Dictionary")
    _write_dictionary(wb, datasets)

    logger.info("Writing Quality_Report")
    _write_quality(wb, datasets)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    try:
        wb.save(str(output_path))
    except PermissionError:
        alt = output_path.with_stem(output_path.stem + "_new")
        wb.save(str(alt))
        output_path = alt
        logger.warning(f"Original file locked; saved to: {alt}")
    logger.info(f"Workbook saved: {output_path}")
    logger.info(f"Total worksheets: {len(wb.sheetnames)}")

    return output_path


if __name__ == "__main__":
    logging.basicConfig(level=logging.INFO, format="%(asctime)s [%(levelname)s] %(message)s")
    print("Export module loaded. Run via main.py.")
